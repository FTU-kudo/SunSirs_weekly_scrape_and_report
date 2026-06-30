"""
Gửi báo cáo giá hàng hoá SunSirs qua Gmail mỗi Chủ Nhật 20:30.

Chạy tự động qua GitHub Actions (send_report.yml).
Yêu cầu 3 GitHub Secrets:
    GMAIL_USER          → địa chỉ Gmail gửi đi
    GMAIL_APP_PASSWORD  → App Password (16 ký tự, không phải mật khẩu Gmail)
    RECIPIENT_EMAILS    → danh sách người nhận, cách nhau bởi dấu phẩy
                          VD: "a@gmail.com,b@ysvn.com.vn,c@ysvn.com.vn"
"""

import os
import smtplib
import logging
import pandas as pd
from io import BytesIO
from pathlib import Path
from datetime import date
from email import encoders
from email.mime.base import MIMEBase
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ──────────────────────────────────────────────────────────────────────────────
# Logging
# ──────────────────────────────────────────────────────────────────────────────
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)-8s  %(message)s",
    datefmt="%H:%M:%S",
)
log = logging.getLogger(__name__)

# ──────────────────────────────────────────────────────────────────────────────
# Đường dẫn
# ──────────────────────────────────────────────────────────────────────────────
BASE_DIR    = Path(__file__).resolve().parent.parent.parent
RAW_PARQUET = BASE_DIR / "data" / "raw" / "commodities" / "sunsirs" / "sunsirs_raw.parquet"

# ──────────────────────────────────────────────────────────────────────────────
# Hàng hoá cần xuất — giữ đồng bộ với sunsirs_scraper.py
# ──────────────────────────────────────────────────────────────────────────────
COMMODITIES_WANTED = {
    "Urea"              : "Urea",
    "Phosphorus yellow" : "Phosphorus yellow",
    "Phosphoric acid"   : "Phosphoric acid",
    "Hydrochloric acid" : "Hydrochloric acid",
    "Sulfuric acid"     : "Sulfuric acid",
}

# ──────────────────────────────────────────────────────────────────────────────
# Helpers
# ──────────────────────────────────────────────────────────────────────────────
def detect_name_col(df: pd.DataFrame) -> str | None:
    for kw in ["tên hàng", "hàng hóa", "commodity", "name"]:
        for col in df.columns:
            if kw.lower() in col.lower():
                return col
    return None


def clean_price(series: pd.Series) -> pd.Series:
    return (
        series
        .str.replace(r"[^\d.\-]", "", regex=True)
        .replace("", pd.NA)
        .pipe(pd.to_numeric, errors="coerce")
    )


# ──────────────────────────────────────────────────────────────────────────────
# Đọc parquet → pivot → build XLSX trong bộ nhớ (không lưu file)
# ──────────────────────────────────────────────────────────────────────────────
def build_xlsx_bytes() -> bytes:
    raw = pd.read_parquet(RAW_PARQUET)
    raw.columns = raw.columns.str.strip()

    name_col  = detect_name_col(raw)
    price_col = next(
        (c for c in raw.columns
         if any(k in c.lower() for k in ["current day price", "price"])),
        None,
    )

    if not name_col or not price_col:
        raise ValueError(f"Không tìm thấy cột tên/giá. Các cột: {list(raw.columns)}")

    # Fuzzy match hàng hoá
    available = raw[name_col].dropna().unique()
    matched   = {}
    for search, label in COMMODITIES_WANTED.items():
        for n in available:
            if search.lower() in n.lower():
                matched[n] = label

    if not matched:
        raise ValueError("Không khớp được hàng hoá nào.")

    # Filter → clean → pivot
    filtered = raw[raw[name_col].isin(matched)].copy()
    filtered["_price"] = clean_price(filtered[price_col])
    filtered["date"]   = pd.to_datetime(filtered["date"], errors="coerce")
    filtered["_label"] = filtered[name_col].map(matched)
    filtered = filtered.dropna(subset=["date", "_price"])

    pivoted = (
        filtered
        .pivot_table(index="date", columns="_label",
                     values="_price", aggfunc="mean")
        .sort_index()
        .reset_index()
    )
    pivoted.columns.name = None
    ordered  = ["date"] + [v for v in COMMODITIES_WANTED.values() if v in pivoted.columns]
    pivoted  = pivoted[ordered]
    n_rows   = len(pivoted)
    n_commod = len(pivoted.columns) - 1

    # ── Build XLSX ──────────────────────────────────────────────────────────
    wb = Workbook()
    ws = wb.active
    ws.title = "Commodity Prices"

    HDR_FILL = PatternFill("solid", fgColor="1F4E79")
    ALT_FILL = PatternFill("solid", fgColor="EBF3FB")
    HDR_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    DAT_FONT = Font(name="Arial", size=10)
    DTE_FONT = Font(name="Arial", size=10, bold=True)
    THIN     = Side(style="thin", color="BDD7EE")
    BDR      = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

    # Header
    headers = ["Date"] + [v for v in COMMODITIES_WANTED.values() if v in pivoted.columns]
    for col_idx, header in enumerate(headers, 1):
        c = ws.cell(row=1, column=col_idx, value=header)
        c.fill      = HDR_FILL
        c.font      = HDR_FONT
        c.border    = BDR
        c.alignment = Alignment(horizontal="center", vertical="center")

    # Data rows
    for row_idx, row_data in enumerate(pivoted.itertuples(index=False), 2):
        fill = ALT_FILL if row_idx % 2 == 0 else PatternFill()
        for col_idx, value in enumerate(row_data, 1):
            c        = ws.cell(row=row_idx, column=col_idx)
            c.border = BDR
            c.fill   = fill
            if col_idx == 1:
                # Cột Date
                c.value         = value.strftime("%Y-%m-%d") if hasattr(value, "strftime") else str(value)
                c.font          = DTE_FONT
                c.alignment     = Alignment(horizontal="center")
            else:
                # Cột giá
                c.value         = round(float(value), 2) if pd.notna(value) else None
                c.font          = DAT_FONT
                c.alignment     = Alignment(horizontal="right")
                c.number_format = "#,##0.00"

    # Info sheet
    ws_info = wb.create_sheet("Info")
    today   = date.today()
    meta    = [
        ("Source",              "sunsirs.com/en"),
        ("Last updated",        str(today)),
        ("Commodities",         ", ".join(COMMODITIES_WANTED.values())),
        ("Rows (trading days)", str(n_rows)),
        ("Generated by",        "YSVN Research – GitHub Actions"),
    ]
    for r, (k, v) in enumerate(meta, 1):
        ws_info.cell(r, 1, k).font = Font(name="Arial", bold=True, size=10)
        ws_info.cell(r, 2, v).font = Font(name="Arial", size=10)
    ws_info.column_dimensions["A"].width = 24
    ws_info.column_dimensions["B"].width = 50

    # Column widths & freeze
    ws.column_dimensions["A"].width = 14
    for col_idx in range(2, n_commod + 2):
        ltr = get_column_letter(col_idx)
        ws.column_dimensions[ltr].width = max(
            len(str(ws.cell(row=1, column=col_idx).value or "")) + 4, 18
        )
    ws.freeze_panes   = "B2"
    ws.row_dimensions[1].height = 24

    # Lưu vào bộ nhớ (không ra file)
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    log.info("✅ XLSX built: %d ngày × %d hàng hoá", n_rows, n_commod)
    return buffer.read()


# ──────────────────────────────────────────────────────────────────────────────
# Gửi email
# ──────────────────────────────────────────────────────────────────────────────
def send_email():
    # Đọc secrets từ biến môi trường (GitHub Actions inject tự động)
    gmail_user     = os.environ["GMAIL_USER"]
    gmail_password = os.environ["GMAIL_APP_PASSWORD"]
    recipients     = [r.strip() for r in os.environ["RECIPIENT_EMAILS"].split(",")]

    today    = date.today()
    filename = f"SunSirs_Commodity_Prices_{today.strftime('%Y-%m-%d')}.xlsx"

    log.info("⏳ Đang build file XLSX…")
    xlsx_bytes = build_xlsx_bytes()

    # ── Soạn email ────────────────────────────────────────────────────────────
    msg              = MIMEMultipart()
    msg["From"]      = gmail_user
    msg["To"]        = ", ".join(recipients)
    msg["Subject"]   = f"[YSVN] Báo cáo giá hàng hoá SunSirs – {today.strftime('%d/%m/%Y')}"

    body = f"""Kính gửi Quý khách hàng,

Đính kèm là báo cáo cập nhật giá hàng hoá từ SunSirs tính đến ngày {today.strftime('%d/%m/%Y')}.

Danh sách hàng hoá:
  • Urea
  • Phosphorus yellow (Phốt pho vàng)
  • Phosphoric acid (Axit phốt pho ric)
  • Hydrochloric acid (Axit clohidric)
  • Sulfuric acid (Axit sunfuric)

Nguồn dữ liệu : sunsirs.com
Tần suất cập nhật: Hàng tuần (Chủ Nhật 20:30)

Trân trọng,
YSVN Research Team
──────────────────────────────────
Vietnam Equity Research | YSVN
""".strip()

    msg.attach(MIMEText(body, "plain", "utf-8"))

    # ── Đính kèm file XLSX ───────────────────────────────────────────────────
    attachment = MIMEBase(
        "application",
        "vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    attachment.set_payload(xlsx_bytes)
    encoders.encode_base64(attachment)
    attachment.add_header(
        "Content-Disposition", "attachment", filename=filename
    )
    msg.attach(attachment)

    # ── Gửi qua Gmail SMTP SSL ───────────────────────────────────────────────
    log.info("⏳ Đang gửi email tới %d người nhận…", len(recipients))
    with smtplib.SMTP_SSL("smtp.gmail.com", 465) as smtp:
        smtp.login(gmail_user, gmail_password)
        smtp.sendmail(gmail_user, recipients, msg.as_string())

    log.info("✅ Email đã gửi thành công!")
    log.info("   Người nhận : %s", ", ".join(recipients))
    log.info("   File đính kèm: %s", filename)


# ──────────────────────────────────────────────────────────────────────────────
# Main
# ──────────────────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    send_email()
